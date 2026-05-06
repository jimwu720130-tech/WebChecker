from __future__ import annotations

import sys
from pathlib import Path


def _to_s(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def main(argv: list[str]) -> int:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        print(f"ERROR: openpyxl not available: {type(e).__name__}: {e}")
        return 2

    if len(argv) < 2:
        print("Usage: python tools/inspect_xlsx.py <xlsx_path> [sheet_name]")
        return 2

    xlsx_path = Path(argv[1])
    if not xlsx_path.exists():
        print(f"ERROR: file not found: {xlsx_path}")
        return 2

    wb = load_workbook(xlsx_path, data_only=True)
    sheet = argv[2] if len(argv) >= 3 else wb.sheetnames[0]
    if sheet not in wb.sheetnames:
        print(f"ERROR: sheet not found: {sheet}; available={wb.sheetnames}")
        return 2

    ws = wb[sheet]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [_to_s(x) for x in header_row]

    flag_keys = ("不符合", "失效", "異常", "錯誤", "無法", "404", "500")
    flagged = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = [_to_s(x) for x in row]
        if not any(vals):
            continue
        blob = " | ".join(vals)
        if any(k in blob for k in flag_keys):
            flagged.append((idx, vals))

    print(f"xlsx={xlsx_path}")
    print(f"sheet={sheet}")
    print(f"headers({len(headers)})={headers}")
    print(f"flagged_rows={len(flagged)} (keys={flag_keys})")

    # print a compact view: row#, url-like column, and any columns containing flags
    url_col = None
    for i, h in enumerate(headers):
        if "網址" in h or "URL" in h.upper():
            url_col = i
            break

    def cols_with_flags(vals: list[str]) -> list[tuple[int, str, str]]:
        out = []
        for i, v in enumerate(vals):
            if any(k in v for k in flag_keys):
                out.append((i, headers[i] if i < len(headers) else f"COL{i+1}", v))
        return out

    for row_i, vals in flagged[:200]:
        u = vals[url_col] if url_col is not None and url_col < len(vals) else ""
        cols = cols_with_flags(vals)
        print(f"- row={row_i} url={u}")
        for i, h, v in cols:
            print(f"  - col={i+1} {h}: {v}")

    if len(flagged) > 200:
        print(f"... truncated; total flagged_rows={len(flagged)}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))

